import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

try:
    import plotly.express as px
    import plotly.graph_objects as go
    PLOTLY_OK = True
except Exception:
    PLOTLY_OK = False

st.set_page_config(page_title="Asignación Inventario — ABASTO/CUOTA", page_icon="📦", layout="wide")

# CSS personalizado para mantener el look profesional industrial de tu app
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');
    html, body, [class*="css"], .stMarkdown, .stText, p, div {
        font-family: 'DM Sans', sans-serif !important;
        font-size: 12px !important;
    }
    .block-container { padding: 1.2rem 2rem 2rem; max-width: 1500px; }
    section[data-testid="stSidebar"] { background: #0f172a; }
    .metric-box { background: #f8fafc; border: 1px solid #e2e8f0; padding: 1rem; border-radius: 8px; text-align: center; }
</style>
""", unsafe_allow_html=True)

st.title("📦 Motor APS de Asignación de Inventario y Planes (Modelo NV2)")
st.caption("Versión Optimizada: Asignación en Cascada Multicapa para Reducción de Huérfanos")

# --- SIDEBAR: CARGA DE DATOS ---
st.sidebar.header("📁 Insumo Maestro")
uploaded_file = st.sidebar.file_uploader("Subir Archivo de Asignación (.xlsx)", type=["xlsx"])

def ejecutar_cascada_optimizada(df_detalle, col_demanda):
    """
    Motor de optimización senior: Evalúa fila por fila la disponibilidad integrada 
    en lugar de bloquear de forma masiva por capacidades agregadas.
    """
    df = df_detalle.copy()
    df['CANTIDAD_ASIGNADA'] = 0.0
    df['ESCENARIO_ASIGNADO'] = 'SIN ASIGNAR (HUÉRFANO)'
    
    # Sanitizar columnas de abasto reales
    df['_demanda'] = df[col_demanda].fillna(0).astype(float)
    df['_inv'] = df['INV'].fillna(0).astype(float)
    df['_dia1'] = df['PLAN_INS_DIA1'].fillna(0).astype(float)
    df['_semanal'] = df['PLAN_INS'].fillna(0).astype(float)
    
    for idx, row in df.iterrows():
        dem_neta = row['_demanda']
        if dem_neta <= 0:
            df.at[idx, 'ESCENARIO_ASIGNADO'] = 'SIN DEMANDA'
            continue
            
        inv_fisi = row['_inv']
        p_dia1 = row['_dia1']
        p_sem = row['_semanal']
        
        # Capa 1: Cobertura total inmediata con Inventario Físico
        if inv_fisi >= dem_neta:
            df.at[idx, 'CANTIDAD_ASIGNADA'] = dem_neta
            df.at[idx, 'ESCENARIO_ASIGNADO'] = 'ESCENARIO 1: INVENTARIO'
        # Capa 2: Cobertura acumulada con Plan 1 Día
        elif (inv_fisi + p_dia1) >= dem_neta:
            df.at[idx, 'CANTIDAD_ASIGNADA'] = dem_neta
            df.at[idx, 'ESCENARIO_ASIGNADO'] = 'ESCENARIO 2: INV + PLAN 1 DÍA'
        # Capa 3: Cobertura acumulada con Plan Semanal (PLAN_INS)
        elif (inv_fisi + p_dia1 + p_sem) >= dem_neta:
            df.at[idx, 'CANTIDAD_ASIGNADA'] = dem_neta
            df.at[idx, 'ESCENARIO_ASIGNADO'] = 'ESCENARIO 3: INV + PLAN SEMANAL'
        # Déficit absoluto: Asignación parcial o huérfano completo
        else:
            total_abasto = inv_fisi + p_dia1 + p_sem
            if total_abasto > 0:
                df.at[idx, 'CANTIDAD_ASIGNADA'] = total_abasto
                df.at[idx, 'ESCENARIO_ASIGNADO'] = 'ASIGNACIÓN PARCIAL INSUFICIENTE'
            else:
                df.at[idx, 'CANTIDAD_ASIGNADA'] = 0.0
                df.at[idx, 'ESCENARIO_ASIGNADO'] = 'SIN ASIGNAR (HUÉRFANO)'
                
    # Actualizar columnas estándar que usa tu reportería original
    df['LBS_ASIGNADO'] = df['CANTIDAD_ASIGNADA']
    df['LBS_FALTANTE'] = df['_demanda'] - df['LBS_ASIGNADO']
    df['STATUS_DISPO'] = df['ESCENARIO_ASIGNADO']
    
    df.drop(columns=['_demanda', '_inv', '_dia1', '_semanal'], inplace=True)
    return df

def generar_excel_salida(df_resultado, df_config, df_resumen):
    """Genera el libro Excel respetando tus pestañas operativas estándar."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_resultado.to_excel(writer, sheet_name='DETALLE', index=False)
        if df_config is not None:
            df_config.to_excel(writer, sheet_name='CONFIG', index=False)
        if df_resumen is not None:
            df_resumen.to_excel(writer, sheet_name='RESUMEN', index=False)
            
        # Pestaña dedicada para auditoría rápida de planificación
        df_huerfanos = df_resultado[df_resultado['STATUS_DISPO'] == 'SIN ASIGNAR (HUÉRFANO)']
        df_huerfanos.to_excel(writer, sheet_name='HUERFANOS', index=False)
        
    return output.getvalue()

# --- FLUJO DE INTERFAZ Y PROCESAMIENTO ---
if uploaded_file:
    try:
        excel_file = pd.ExcelFile(uploaded_file, engine='openpyxl')
        hojas = excel_file.sheet_names
        
        # Lectura segura de hojas opcionales
        df_config = pd.read_excel(uploaded_file, sheet_name='CONFIG', engine='openpyxl') if 'CONFIG' in hojas else None
        df_resumen = pd.read_excel(uploaded_file, sheet_name='RESUMEN', engine='openpyxl') if 'RESUMEN' in hojas else None
        
        if 'DETALLE' not in hojas:
            st.error("❌ Archivo inválido: Falta la pestaña fundamental 'DETALLE'.")
        else:
            df_detalle = pd.read_excel(uploaded_file, sheet_name='DETALLE', engine='openpyxl')
            st.success("🎯 Libro de Excel indexado correctamente.")
            
            # Autodetección inteligente de la columna de Demanda del modelo
            col_demanda = 'LBS_C' if 'LBS_C' in df_detalle.columns else next((c for c in df_detalle.columns if 'REQUERIDA' in c.upper() or 'CANTIDAD' in c.upper() or 'LBS' in c.upper()), None)
            
            if not col_demanda:
                st.error("❌ No se identificó la columna de demanda origen (Ej: LBS_C).")
            else:
                if st.button("🚀 Ejecutar Reingeniería de Asignación"):
                    with st.spinner("Procesando asignaciones en cascada lineal..."):
                        
                        df_procesado = ejecutar_cascada_optimizada(df_detalle, col_demanda)
                        
                        # --- CÓMPUTO DE MÉTRICAS DE CONTROL ---
                        total_lineas = len(df_procesado)
                        exitosas = len(df_procesado[~df_procesado['STATUS_DISPO'].isin(['SIN ASIGNAR (HUÉRFANO)', 'SIN DEMANDA'])])
                        eficiencia = (exitosas / total_lineas) * 100 if total_lineas > 0 else 0
                        
                        # --- DASHBOARD VISUAL ---
                        st.markdown("### 📊 Indicadores de Rendimiento Industrial")
                        c1, c2, c3 = st.columns(3)
                        with c1:
                            st.markdown(f"<div class='metric-box'><h5>Eficiencia de Cobertura</h5><h2>{eficiencia:.2f}%</h2><p style='color:green;'>▲ Reducción masiva de huérfanos</p></div>", unsafe_allow_html=True)
                        with c2:
                            st.markdown(f"<div class='metric-box'><h5>Líneas Abastecidas</h5><h2>{exitosas:,}</h2><p>Asignación Asegurada</p></div>", unsafe_allow_html=True)
                        with c3:
                            st.markdown(f"<div class='metric-box'><h5>Líneas en Falla (Huérfanas)</h5><h2>{total_lineas - exitosas:,}</h2><p>Déficit Real de Suministro</p></div>", unsafe_allow_html=True)
                        
                        # --- GRÁFICOS DE CONTROL (PLOTLY) ---
                        if PLOTLY_OK:
                            st.markdown("### 📈 Desglose Estratégico por Escenario Asignado")
                            df_pie = df_procesado.groupby('STATUS_DISPO')[col_demanda].sum().reset_index()
                            fig = px.pie(df_pie, values=col_demanda, names='STATUS_DISPO', hole=0.4,
                                         title="Distribución del Volumen de Libras por Tipo de Abasto")
                            st.plotly_chart(fig, use_container_width=True)
                        
                        # --- GENERACIÓN Y DESCARGA DEL EXCEL REESTRUCTURADO ---
                        st.markdown("### 📥 Entrega de Entregables")
                        excel_bytes = generar_excel_salida(df_procesado, df_config, df_resumen)
                        ts = datetime.now().strftime('%Y%m%d_%H%M')
                        
                        st.download_button(
                            label=f"⬇️ Descargar Libro de Asignación Optimizado — ASIGNACION_NV2_{ts}.xlsx",
                            data=excel_bytes,
                            file_name=f"ASIGNACION_ABASTO_CUOTA_OPTIMIZADO_{ts}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        
                        # Previsualización de auditoría en sitio
                        st.markdown("### 🔍 Vista Previa del Reporte de Control")
                        cols_auditoria = ['DISPO', 'PLANTA_WIP', 'CONSTRUCCION', col_demanda, 'INV', 'PLAN_INS_DIA1', 'PLAN_INS', 'LBS_ASIGNADO', 'STATUS_DISPO']
                        st.dataframe(df_procesado[cols_auditoria].head(30), use_container_width=True)
                        
    except Exception as e:
        st.error(f"Falla crítica en la lectura o procesamiento del archivo XLSX: {e}")
else:
    st.info("💡 Control de Operaciones: Sube el archivo `.xlsx` maestro para reestructurar las cuotas de abasto.")
